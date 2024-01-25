import openpyxl
import json

# Load the Excel file
workbook = openpyxl.load_workbook('test.xlsx')

states = {
    "Alabama": {"X":(-92,-88), "Y":(30,35), "avgs":[]},
    "Alaska": {"X":(-165,-140), "Y":(55,70), "avgs":[]},
    "Arizona": {"X":(-115,-109), "Y":(31,37), "avgs":[]},
    "Arkansas": {"X":(-95,-90), "Y":(33,37), "avgs":[]},
    "California": {"X":(-125,-115), "Y":(32,42), "avgs":[]},
    "Colorado": {"X":(-109,-102), "Y":(37,41), "avgs":[]},
    "Connecticut": {"X":(-74,-72), "Y":(41,42), "avgs":[]},
    "Delaware": {"X":(-76,-75), "Y":(38,40), "avgs":[]},
    "Florida": {"X":(-88,-80), "Y":(25,31), "avgs":[]},
    "Georgia": {"X":(-86,-81), "Y":(30,35), "avgs":[]},
    "Hawaii": {"X":(-160,-155), "Y":(15,25), "avgs":[]},
    "Idaho": {"X":(-117,-111), "Y":(42,49), "avgs":[]},
    "Illinois": {"X":(-92,-87), "Y":(37,43), "avgs":[]},
    "Indiana": {"X":(-88,-85), "Y":(38,42), "avgs":[]},
    "Iowa": {"X":(-97,-90), "Y":(40,44), "avgs":[]},
    "Kansas": {"X":(-102,-95), "Y":(37,40), "avgs":[]},
    "Kentucky": {"X":(-90,-82), "Y":(37,39), "avgs":[]},
    "Louisiana": {"X":(-94,-90), "Y":(29,33), "avgs":[]},
    "Maine": {"X":(-71,-67), "Y":(43,48), "avgs":[]},
    "Maryland": {"X":(-79,-75), "Y":(38,40), "avgs":[]},
    "Massachusetts": {"X":(-74,-70), "Y":(41,43), "avgs":[]},
    "Michican": {"X":(-91,-82), "Y":(42,47), "avgs":[]},
    "Minessota": {"X":(-97,-89), "Y":(43,49), "avgs":[]},
    "Mississippi": {"X":(-94,-91), "Y":(30,35), "avgs":[]},
    "Missouri": {"X":(-96,-89), "Y":(36,41), "avgs":[]},
    "Montana": {"X":(-116,-104), "Y":(45,49), "avgs":[]},
    "Nebraska": {"X":(-104,-95), "Y":(40,43), "avgs":[]},
    "Nevada": {"X":(-120,-114), "Y":(35,42), "avgs":[]},
    "New Hampshire": {"X":(-73,-70), "Y":(42,45), "avgs":[]},
    "New Jersey": {"X":(-76,-74), "Y":(39,42), "avgs":[]},
    "New Mexico": {"X":(-109,-103), "Y":(32,37), "avgs":[]},
    "New York": {"X":(-80,-73), "Y":(41,45), "avgs":[]},
    "North Carolina": {"X":(-85,-76), "Y":(34,37), "avgs":[]},
    "North Dakota": {"X":(-104,-97), "Y":(46,49), "avgs":[]},
    "Ohio": {"X":(-85,-80), "Y":(38,42), "avgs":[]},
    "Oklahoma": {"X":(-103,-94), "Y":(34,37), "avgs":[]},
    "Oregon": {"X":(-125,-117), "Y":(42,46), "avgs":[]},
    "Pennsylvania": {"X":(-81,-75), "Y":(40,42), "avgs":[]},
    "Rhode Island": {"X":(-72,-71), "Y":(41,42), "avgs":[]},
    "South Carolina": {"X":(-83,-79), "Y":(32,35), "avgs":[]},
    "South Dakota": {"X":(-104,-96), "Y":(43,46), "avgs":[]},
    "Tennessee": {"X":(-90,-82), "Y":(35,37), "avgs":[]},
    "Texas": {"X":(-107,-94), "Y":(26,37), "avgs":[]},
    "Utah": {"X":(-115,-109), "Y":(37,42), "avgs":[]},
    "Vermont": {"X":(-73,-71), "Y":(43,45), "avgs":[]},
    "Virginia": {"X":(-84,-75), "Y":(36,40), "avgs":[]},
    "Washington": {"X":(-125,-117), "Y":(45,50), "avgs":[]},
    "West Virginia": {"X":(-83,-78), "Y":(37,40), "avgs":[]},
    "Wisconsin": {"X":(-93,-87), "Y":(42,47), "avgs":[]},
    "Wyoming": {"X":(-111,-104), "Y":(41,45), "avgs":[]}
}


# Loop through the sheets
sheet_names = ["{}-22".format(i) for i in range(1, 13)]
for s in states:
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        # Grab the state's coordinate bounds
        lower_row = states[s]["Y"][0]
        upper_row = states[s]["Y"][1]
        lower_col = states[s]["X"][0]
        upper_col = states[s]["X"][1]

        # Find start and end rows based on the state's bounds
        start_row = None
        end_row = None
        for i, row in enumerate(sheet.iter_rows()):
            if start_row is None and row[0].value is not None and row[0].value >= lower_row:
                start_row = i + 1
            if end_row is None and row[0].value is not None and row[0].value > upper_row:
                end_row = i
                break
        if start_row is None:
            start_row = 1
        if end_row is None:
            end_row = sheet.max_row

        # Find start and end columns based on the state's bounds
        start_col = None
        end_col = None
        for j, col in enumerate(sheet.iter_cols()):
            if start_col is None and col[0].value is not None and col[0].value >= lower_col:
                start_col = openpyxl.utils.get_column_letter(col[0].column)
            if end_col is None and col[0].value is not None and col[0].value > upper_col:
                end_col = openpyxl.utils.get_column_letter(col[0].column)
                break
        if start_col is None:
            start_col = 'A'
        if end_col is None:
            end_col = openpyxl.utils.get_column_letter(sheet.max_column)

        # Extract data within specified range
        # (ex, in) for rows, (ex, ex) for columns
        all_vals = []
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=openpyxl.utils.column_index_from_string(start_col), max_col=openpyxl.utils.column_index_from_string(end_col)):
            for cell in row:
                all_vals.append(cell.value)

        # add averages to states dictionary
        states[s]["avgs"].append(sum(all_vals)/len(all_vals))


# Convert everything from Kelvin to Farenheit 
for s in states:
    states[s]["avgs"] = [round((t - 273.15) * 9/5 + 32, 2) for t in states[s]["avgs"]]



# TURN THIS INTO A JSON


state_data = []

# Loop through the input data and create a dictionary for each state
for s in states:
    state_dict = {
        "name": s,
        "data": states[s]["avgs"]
    }
    state_data.append(state_dict)

# Create the final dictionary with the list of state data dictionaries
output_data = {
    "_id":"monthlyStAvg",
    "states": state_data
}

# Convert the dictionary to JSON and print it
json_data = json.dumps(output_data, indent=4)


with open("output.json", "w") as outfile:
    outfile.write(json_data)